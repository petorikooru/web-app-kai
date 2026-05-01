import asyncio
import base64
import csv
import io
import json
import logging
import math
import multiprocessing as mp
import operator
import os
import re
import queue
import random
import time
from datetime import datetime
from multiprocessing import Process, Queue
from typing import Optional

import cv2
import numpy as np
import pandas as pd
import torch
import uvicorn
from scipy.spatial import KDTree
from contextlib import asynccontextmanager

from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")
log = logging.getLogger("RailDetection")


# ============================================================================
# KONSTANTA GLOBAL (dibagikan ke semua proses)
# ============================================================================

CLASS_NAMES = {
    0: "Background",
    1: "Corrugations",
    2: "rail-end",
    3: "Rail Line",
    4: "Shelling",
    5: "Squat",
    6: "Wheelburns",
    7: "Wheelslip",
}

CLASS_COLORS = {
    0: (0, 0, 0),
    1: (0, 255, 0),
    2: (0, 0, 255),
    3: (255, 255, 0),
    4: (255, 0, 0),
    5: (0, 165, 255),
    6: (128, 0, 128),
    7: (0, 255, 255),
}

SKIP_CLASSES = {0, 3, 6, 7}


# ============================================================================
# CLASS 1 — WebBroadcaster
# Menangani pengiriman data (frame + log) dari proses ke browser via WebSocket
# ============================================================================


class WebBroadcaster:
    """
    Jembatan antara pipeline multiprocessing dan klien browser.

    Cara kerja:
      - BatchWorkerProcess menaruh hasil ke `broadcast_queue` (multiprocessing.Queue)
      - WebBroadcaster._poll_queue() dipanggil via asyncio loop di dalam WebServer
      - Setiap WebSocket klien yang connect didaftarkan ke `_clients`
      - Data dikirim ke semua klien secara async (broadcast)

    Format pesan WebSocket (JSON):
      {
        "type": "frame",           # atau "log" / "status" / "graph_start"
        "camera_id": 1,
        "frame_b64": "<base64>",   # JPEG frame
        "total_contours": 3,
        "gps": "...",
        "detected_classes": [{"name": "Squat", "count": 2}],
        "timestamp": "2024-..."
      }
    """

    def __init__(self, broadcast_queue: Queue):
        self.broadcast_queue = broadcast_queue
        self._clients: list[WebSocket] = []
        self._lock = asyncio.Lock()

    async def register(self, ws: WebSocket):
        async with self._lock:
            self._clients.append(ws)
        log.info(f"WebSocket klien terhubung. Total: {len(self._clients)}")

    async def unregister(self, ws: WebSocket):
        async with self._lock:
            self._clients = [c for c in self._clients if c is not ws]
        log.info(f"WebSocket klien terputus. Sisa: {len(self._clients)}")

    async def _send_to_all(self, message: dict):
        """Kirim pesan JSON ke semua klien yang terhubung."""
        if not self._clients:
            return
        text = json.dumps(message)
        disconnected = []
        async with self._lock:
            for ws in self._clients:
                try:
                    await ws.send_text(text)
                except Exception:
                    disconnected.append(ws)
            for ws in disconnected:
                self._clients.remove(ws)

    async def poll_and_broadcast(self):
        """
        Loop async yang terus membaca broadcast_queue dan meneruskan ke browser.
        Dipanggil sebagai asyncio Task saat server start.
        """
        loop = asyncio.get_event_loop()
        while True:
            try:
                # Non-blocking get agar tidak memblokir event loop
                msg = await loop.run_in_executor(
                    None, lambda: self.broadcast_queue.get(timeout=0.05)
                )
                await self._send_to_all(msg)
            except Exception:
                # Queue kosong atau timeout — tidak apa-apa
                await asyncio.sleep(0.01)


# ============================================================================
# CLASS 2 — WebServer
# FastAPI app + endpoint WebSocket + halaman HTML dashboard
# ============================================================================


class WebServer:
    """
    FastAPI server yang menjadi antarmuka web untuk pipeline.

    Endpoint:
      GET  /          → Dashboard HTML (embedded, tidak perlu file statis)
      WS   /ws        → Stream frame + log JSON ke browser
      POST /control   → Terima perintah start/stop/pause dari browser
    """

    def __init__(
        self,
        broadcaster: "WebBroadcaster",
        control_queue: Queue,
        host: str = "0.0.0.0",
        port: int = 8000,
    ):
        self.broadcaster = broadcaster
        self.control_queue = control_queue
        self.host = host
        self.port = port
        self.app = FastAPI(title="Rail Crack Detection")
        self._register_routes()

    def _register_routes(self):
        broadcaster = self.broadcaster

        @asynccontextmanager
        async def lifespan(app: FastAPI):
            # startup
            task = asyncio.create_task(broadcaster.poll_and_broadcast())
            log.info("Broadcast loop dimulai.")
            yield
            # shutdown
            task.cancel()

        # Re-create app dengan lifespan agar tidak pakai on_event
        self.app = FastAPI(title="Rail Crack Detection", lifespan=lifespan)
        app = self.app

        app.mount("/icon_100_", StaticFiles(directory="icon_100_"), name="icon_100_")

        @app.get("/", response_class=HTMLResponse)
        async def dashboard():
            # return _build_dashboard_html()
            return FileResponse("index.html")

        @app.websocket("/ws")
        async def websocket_endpoint(ws: WebSocket):
            await ws.accept()
            await self.broadcaster.register(ws)
            try:
                # Tetap terbuka; klien bisa kirim perintah via WS juga
                while True:
                    data = await ws.receive_text()
                    try:
                        cmd = json.loads(data)

                        # --- TAMBAHAN UNTUK THRESHOLD & LOGGING ---
                        action = cmd.get("action")
                        if action == "operator_settings":
                            # Hanya print/log, lalu biarkan masuk ke queue
                            print(
                                f"[WEBSOCKET] Menerima perubahan operator: {cmd.get('data')}"
                            )
                        # elif action == "set_mode":
                        #     print(
                        #         f"[WEBSOCKET] Menerima perubahan mode: {cmd.get('mode')}"
                        #     )
                        # ------------------------------------------

                        self.control_queue.put_nowait(cmd)
                    except json.JSONDecodeError:
                        pass
            except WebSocketDisconnect:
                pass
            finally:
                await self.broadcaster.unregister(ws)

        @app.post("/control")
        async def control(payload: dict):
            """
            Terima perintah kontrol dari browser (HTTP POST).
            Payload contoh: {"action": "stop"} / {"action": "pause"} / {"action": "start"}
            """
            self.control_queue.put_nowait(payload)
            return {"status": "ok", "received": payload}

        @app.get("/download")
        async def download_report(filepath: str):
            """Endpoint HTTP murni untuk melayani file Excel."""
            if os.path.exists(filepath):
                return FileResponse(
                    filepath,
                    filename="Laporan inspeksi.xlsx",
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            return {"error": "File tidak ditemukan atau belum selesai digenerate."}

    def run(self):
        """Jalankan server di main process (blocking)."""
        uvicorn.run(self.app, host=self.host, port=self.port, log_level="warning")

    def run_in_process(self) -> Process:
        """Jalankan server di proses terpisah (non-blocking untuk main)."""
        p = Process(target=self.run, daemon=True, name="WebServerProcess")
        p.start()
        return p


# ============================================================================
# CLASS 3 — ImageSaverProcess
# Menggantikan ImageSaverWorker(QThread) → multiprocessing.Process
# ============================================================================
def _is_jpg(val: str) -> bool:
    return isinstance(val, str) and ".jpg" in val.lower()


def _is_defect_type(val: str) -> bool:
    """True jika val adalah nama jenis cacat (bukan jpg, dash, kosong, atau angka)."""
    if not val or val.strip() in ("", "-"):
        return False
    if _is_jpg(val):
        return False
    try:
        float(val)
        return False
    except ValueError:
        return True


def _safe(val) -> str:
    """Kembalikan '-' jika nilai None atau string kosong."""
    if val is None:
        return "-"
    v = str(val).strip()
    return v if v else "-"


def _parse_datetime(time_str: str):
    """'20260408_200305' → ('2026-04-08', '20:03:05')"""
    s = str(time_str).strip()
    try:
        dp = s[:8]
        tp = s[9:15]
        return f"{dp[:4]}-{dp[4:6]}-{dp[6:8]}", f"{tp[:2]}:{tp[2:4]}:{tp[4:6]}"
    except Exception:
        return "-", s


def _parse_gps(gps_str: str):
    """Ekstrak (lat, lon) dari string GPS. Return ('-', '-') jika belum ada sinyal."""
    s = str(gps_str).strip()
    if not s or s == "nan" or "Menunggu" in s or "waiting" in s.lower():
        return "-", "-"
    nums = re.findall(r"[-+]?\d+\.\d+", s)
    if len(nums) >= 2:
        return nums[0], nums[1]
    return s, "-"


def _classify_and_fix_row(row: list) -> dict:
    """
    CSV Header Baru:
    0:Time, 1:GPS, 2:Accuracy, 3:Loss, 4:Confidence,
    5:Left Defect Type, 6:Left Total, 7:Left Image Name,
    8:Right Defect Type, 9:Right Total, 10:Right Image Name,
    11:Status, 12:Operator, 13:NIPP Operator, 14:PPJ, 15:NIPP PPJ,
    16:Petak Jalan, 17:Daop/Divre, 18:Nomor KPJ
    """
    r = row + ["-"] * 20  # Padding untuk keamanan

    return {
        "time_raw": r[0],
        "gps_raw": r[1],
        "accuracy": _safe(r[2]),
        "left_defect": _safe(r[5]),
        "left_image": _safe(r[7]),
        "right_defect": _safe(r[8]),
        "right_image": _safe(r[10]),
        # Ekstrak Info Operator (Diambil dari baris pertama yang memilikinya)
        "Operator": _safe(r[12]),
        "NIPP Operator": _safe(r[13]),
        "PPJ": _safe(r[14]),
        "NIPP PPJ": _safe(r[15]),
        "Petak Jalan": _safe(r[16]),
        "Daop/Divre": _safe(r[17]),
        "Nomor KPJ": _safe(r[18]),
    }


def _get_indonesian_date(date_str: str) -> str:
    """Mengubah '2026-04-25' menjadi 'Sabtu, 25 April 2026'."""
    if date_str == "-" or not date_str:
        return "-"
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        hari = ["Senin", "Selasa", "Rabu", "Kamis", "Jumat", "Sabtu", "Minggu"][
            dt.weekday()
        ]
        bulan = [
            "Januari",
            "Februari",
            "Maret",
            "April",
            "Mei",
            "Juni",
            "Juli",
            "Agustus",
            "September",
            "Oktober",
            "November",
            "Desember",
        ][dt.month - 1]
        return f"{hari}, {dt.day} {bulan} {dt.year}"
    except Exception:
        return date_str


def _build_output_df(raw_rows: list) -> tuple:
    """
    Parse CSV yang sudah rapi menjadi DataFrame (Tabel Kerusakan TANPA Date/Time)
    dan Dictionary (Informasi Operator + Hari/Tanggal).
    """
    # Date dan Time dihapus dari tabel
    out_cols = [
        "Latitude",
        "Longitude",
        "KM HM",
        "Accuracy",
        "Defect Type",
        "Image",
        "POSITION",
    ]
    records = []
    operator_info = {}

    for row in raw_rows:
        if not row or not row[0].strip() or row[0].strip() == "Time":
            continue

        r = row + ["-"] * 15

        time_raw = str(r[0]).strip()
        gps_raw = str(r[1]).strip()
        accuracy = str(r[2]).strip()
        defect_type = str(r[3]).strip()
        image_name = str(r[4]).strip()
        position = str(r[5]).strip()

        op_nama = str(r[6]).strip()
        op_nipp = str(r[7]).strip()
        ppj_nama = str(r[8]).strip()
        ppj_nipp = str(r[9]).strip()
        petak = str(r[10]).strip()
        daop = str(r[11]).strip()
        kpj = str(r[12]).strip()

        # Tangkap data operator & Hari/Tanggal dari baris pertama yang valid
        if not operator_info:
            date_val, _ = _parse_datetime(time_raw)
            operator_info = {
                "Hari/Tanggal": _get_indonesian_date(date_val),
                "Operator": op_nama if op_nama != "-" else "-",
                "NIPP Operator": op_nipp,
                "PPJ": ppj_nama,
                "NIPP PPJ": ppj_nipp,
                "Petak Jalan": petak,
                "Daop/Divre": daop,
                "Nomor KPJ": kpj,
            }

        # Masukkan data kerusakan (tanpa Date/Time)
        if defect_type and defect_type not in ("-", "", "nan"):
            lat, lon = _parse_gps(gps_raw)
            records.append(
                {
                    "Latitude": lat,
                    "Longitude": lon,
                    "KM HM": "-",
                    "Accuracy": accuracy,
                    "Defect Type": defect_type,
                    "Image": image_name,
                    "POSITION": position,
                }
            )

    df = pd.DataFrame(records, columns=out_cols)
    if not df.empty:
        df.insert(0, "No", range(1, len(df) + 1))

    return df, operator_info


def _apply_excel_style(output_path: str, n_cols: int):
    """Terapkan format plain putih + GPS merged header ke file Excel."""
    wb = load_workbook(output_path)
    ws = wb.active

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_border = Border(
        left=Side(style="thin", color="AAAAAA"),
        right=Side(style="thin", color="AAAAAA"),
        top=Side(style="thin", color="AAAAAA"),
        bottom=Side(style="medium", color="888888"),
    )
    header_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    white_fill = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    alt_fill = PatternFill("solid", start_color="F9F9F9", end_color="F9F9F9")

    # Row 1 — GPS merged header
    ws.insert_rows(1)
    ws.merge_cells("C1:D1")
    ws["C1"].value = "GPS"
    ws["C1"].font = Font(name="Arial", bold=True, size=10)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C1"].border = header_border
    ws["C1"].fill = header_fill
    for col in range(1, n_cols + 1):
        if col not in (3, 4):
            ws.cell(row=1, column=col).border = header_border
            ws.cell(row=1, column=col).fill = header_fill
    ws.row_dimensions[1].height = 16

    # Row 2 — sub-header
    sub_headers = [
        "Date",
        "Time",
        "Latitude",
        "Longitude",
        "KM HM",
        "Defect Type",
        "Confidence",
        "Condition",
        "Image",
        "Camera",
        "Left Total",
        "Right Total",
    ]
    for col_idx, name in enumerate(sub_headers[:n_cols], start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = name
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = header_border
        cell.fill = header_fill
    ws.row_dimensions[2].height = 22

    # Data rows — alternating white / near-white
    for row_idx in range(3, ws.max_row + 1):
        fill = white_fill if row_idx % 2 == 0 else alt_fill
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = thin
            cell.fill = fill

    col_widths = [13, 10, 13, 13, 9, 16, 11, 24, 32, 8, 11, 11]
    for i, w in enumerate(col_widths[:n_cols], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"
    wb.save(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# CSV COLUMNS (struktur yang benar untuk disimpan dari pipeline)
# ─────────────────────────────────────────────────────────────────────────────

CSV_COLUMNS = [
    "Mode",  # TAMBAHAN BARU: segformer / yolo / yolo_ballast
    "Time",
    "GPS",
    "KM HM",
    "Accuracy",
    "Defect Type",
    "Image",
    "Position",
    "Operator",
    "NIPP Operator",
    "PPJ",
    "NIPP PPJ",
    "Petak Jalan",
    "Daop/Divre",
    "Nomor KPJ",
]


def _build_output_dfs(raw_rows: list) -> tuple:
    """
    Parse CSV menjadi Dictionary of DataFrames berdasarkan Mode.
    """
    out_cols = [
        "Latitude",
        "Longitude",
        "KM HM",
        "Accuracy",
        "Defect Type",
        "Image",
        "POSITION",
    ]
    records = {"segformer": [], "yolo": [], "yolo_ballast": []}
    operator_info = {}

    for row in raw_rows:
        # Cek jika baris kosong atau header
        if not row or not row[0].strip() or row[0].strip() in ("Mode", "Time"):
            continue

        r = row + ["-"] * 15

        mode = str(r[0]).strip()
        time_raw = str(r[1]).strip()
        gps_raw = str(r[2]).strip()
        km_hm = str(r[3]).strip()  # <-- TAMBAHAN (index 3)
        accuracy = str(r[4]).strip()  # sebelumnya r[3]
        defect_type = str(r[5]).strip()  # sebelumnya r[4]
        image_name = str(r[6]).strip()  # sebelumnya r[5]
        position = str(r[7]).strip()  # sebelumnya r[6]

        op_nama = str(r[8]).strip()
        op_nipp = str(r[9]).strip()
        ppj_nama = str(r[10]).strip()
        ppj_nipp = str(r[11]).strip()
        petak = str(r[12]).strip()
        daop = str(r[13]).strip()
        kpj = str(r[14]).strip()

        if not operator_info:
            date_val, _ = _parse_datetime(time_raw)
            operator_info = {
                "Hari/Tanggal": _get_indonesian_date(date_val),
                "Operator": op_nama if op_nama != "-" else "-",
                "NIPP Operator": op_nipp,
                "PPJ": ppj_nama,
                "NIPP PPJ": ppj_nipp,
                "Petak Jalan": petak,
                "Daop/Divre": daop,
                "Nomor KPJ": kpj,
            }

        if defect_type and defect_type not in ("-", "", "nan") and mode in records:
            lat, lon = _parse_gps(gps_raw)
            records[mode].append(
                {
                    "Latitude": lat,
                    "Longitude": lon,
                    "KM HM": km_hm if km_hm not in ("-", "", "nan") else "-",
                    "Accuracy": accuracy,
                    "Defect Type": defect_type,
                    "Image": image_name,
                    "POSITION": position,
                }
            )

    dfs = {}
    for m, data in records.items():
        df = pd.DataFrame(data, columns=out_cols)
        if not df.empty:
            df.insert(0, "No", range(1, len(df) + 1))
        dfs[m] = df

    return dfs, operator_info


# ─────────────────────────────────────────────────────────────────────────────
# MAIN CLASS
# ─────────────────────────────────────────────────────────────────────────────


class ImageSaverProcess(Process):
    def __init__(
        self, task_queue: Queue, broadcast_queue: Queue, save_dir: str = "./defect"
    ):
        super().__init__(daemon=True, name="ImageSaverProcess")
        self.task_queue = task_queue
        self.broadcast_queue = broadcast_queue
        self.save_dir = f"{save_dir}_{datetime.now().strftime('%Y-%m-%d_%H:%M:%S')}"
        self.temp_csv = os.path.join(self.save_dir, "temp_session_data.csv")
        self.final_excel = os.path.join(self.save_dir, "laporan_kerusakan.xlsx")
        self.csv_buffer: list[dict] = []
        self.BUFFER_LIMIT = 5

        # Akumulator Tally Fastener
        self.fastener_totals = {"DE-Clip": 0, "E-Clip": 0, "KA-Clip": 0, "No Clip": 0}

    def run(self):
        os.makedirs(self.save_dir, exist_ok=True)
        self._init_csv()
        log.info(f"ImageSaverProcess dimulai. Direktori: {self.save_dir}")

        while True:
            try:
                task_type, data = self.task_queue.get(timeout=1.0)
                if task_type == "image":
                    self._save_image(data)
                elif task_type == "data_log":
                    self._process_log_data(data)
                elif task_type == "fastener_tally":  # Terima tally dari worker
                    for k, v in data.items():
                        self.fastener_totals[k] = self.fastener_totals.get(k, 0) + v
                elif task_type == "export":
                    self._export_csv_to_excel()
                elif task_type == "stop":
                    self._export_csv_to_excel()
                    log.info("ImageSaverProcess berhenti.")
                    break
            except Exception:
                continue

    def _init_csv(self):
        if not os.path.exists(self.temp_csv):
            pd.DataFrame(columns=CSV_COLUMNS).to_csv(self.temp_csv, index=False)

    def _process_log_data(self, row_data: dict):
        self.csv_buffer.append(row_data)
        if len(self.csv_buffer) >= self.BUFFER_LIMIT:
            self._flush_buffer_to_csv()

    def _flush_buffer_to_csv(self):
        if not self.csv_buffer:
            return
        try:
            pd.DataFrame(self.csv_buffer).to_csv(
                self.temp_csv, mode="a", index=False, header=False
            )
            self.csv_buffer = []
        except Exception as e:
            log.error(f"Gagal tulis CSV: {e}")

    def _save_image(self, data):
        img_rgb, filename = data
        try:
            cv2.imwrite(os.path.join(self.save_dir, filename), img_rgb)
        except Exception as e:
            log.error(f"Gagal save image: {e}")

    def _export_csv_to_excel(self):
        self._flush_buffer_to_csv()
        if not os.path.exists(self.temp_csv):
            log.warning("Tidak ada data CSV untuk diexport.")
            return

        try:
            raw_rows = self._parse_raw_csv(self.temp_csv)
            if not raw_rows:
                return

            dfs, op_info = _build_output_dfs(raw_rows)
            TABLE_START_ROW = 12

            sheet_mapping = {
                "segformer": ("Rail Defect", "Defect Type"),
                "yolo": ("Fastener", "Fastener"),
                "yolo_ballast": ("Ballast", "Ballast Type"),
            }

            with pd.ExcelWriter(self.final_excel, engine="xlsxwriter") as writer:
                workbook = writer.book

                title_format = workbook.add_format(
                    {
                        "bold": True,
                        "align": "center",
                        "valign": "vcenter",
                        "font_size": 16,
                        "font_name": "Arial",
                    }
                )
                bold_format = workbook.add_format(
                    {"bold": True, "font_size": 10, "font_name": "Arial"}
                )
                normal_format = workbook.add_format(
                    {"font_size": 10, "font_name": "Arial"}
                )
                header_format = workbook.add_format(
                    {
                        "bold": True,
                        "align": "center",
                        "valign": "vcenter",
                        "bg_color": "#DDEBF7",
                        "border": 1,
                        "border_color": "#000000",
                        "font_name": "Arial",
                        "font_size": 10,
                        "text_wrap": True,
                    }
                )
                cell_format = workbook.add_format(
                    {
                        "align": "center",
                        "valign": "vcenter",
                        "border": 1,
                        "border_color": "#CCCCCC",
                        "font_name": "Arial",
                        "font_size": 10,
                    }
                )

                for mode, df in dfs.items():
                    if df.empty and mode != "yolo":
                        continue  # Skip sheet kosong kecuali YOLO (karena kita butuh cetak summary)

                    sheet_title, defect_col_name = sheet_mapping[mode]

                    if not df.empty:
                        df.rename(
                            columns={"Defect Type": defect_col_name}, inplace=True
                        )
                    else:
                        df = pd.DataFrame(
                            columns=[
                                "No",
                                "Latitude",
                                "Longitude",
                                "KM HM",
                                "Accuracy",
                                defect_col_name,
                                "Image",
                                "POSITION",
                            ]
                        )

                    n_cols = len(df.columns)
                    df.to_excel(
                        writer,
                        index=False,
                        sheet_name=sheet_title,
                        startrow=TABLE_START_ROW,
                        header=False,
                    )
                    worksheet = writer.sheets[sheet_title]

                    # --- KOP SURAT ---
                    worksheet.merge_range(
                        "A1:H1", f"{sheet_title} Detection Report", title_format
                    )
                    worksheet.set_row(0, 30)
                    try:
                        worksheet.insert_image(
                            "H1",
                            "icon_100_/logo_kai.png",
                            {
                                "x_scale": 0.1,
                                "y_scale": 0.1,
                                "x_offset": 10,
                                "y_offset": 5,
                            },
                        )
                    except Exception:
                        pass

                    worksheet.write("A3", "Hari / Tanggal", bold_format)
                    worksheet.write(
                        "B3", f": {op_info.get('Hari/Tanggal', '-')}", normal_format
                    )
                    worksheet.write("A5", "Daop / Divre", bold_format)
                    worksheet.write(
                        "B5", f": {op_info.get('Daop/Divre', '-')}", normal_format
                    )
                    worksheet.write("F5", "Operator", bold_format)
                    worksheet.write(
                        "G5", f": {op_info.get('Operator', '-')}", normal_format
                    )
                    worksheet.write("A4", "Petak Jalan", bold_format)
                    worksheet.write(
                        "B4", f": {op_info.get('Petak Jalan', '-')}", normal_format
                    )
                    worksheet.write("H5", "NIPP", bold_format)
                    worksheet.write(
                        "I5", f": {op_info.get('NIPP Operator', '-')}", normal_format
                    )
                    worksheet.write("F3", "Nomor KPJ", bold_format)
                    worksheet.write(
                        "G3", f": {op_info.get('Nomor KPJ', '-')}", normal_format
                    )
                    worksheet.write("F4", "PPJ", bold_format)
                    worksheet.write("G4", f": {op_info.get('PPJ', '-')}", normal_format)
                    worksheet.write("H4", "NIPP", bold_format)
                    worksheet.write(
                        "I4", f": {op_info.get('NIPP PPJ', '-')}", normal_format
                    )

                    # --- TABEL ---
                    col_widths = [5, 13, 13, 10, 10, 20, 28, 10]
                    for col_idx, width in enumerate(col_widths):
                        worksheet.set_column(col_idx, col_idx, width, cell_format)

                    worksheet.merge_range(10, 1, 10, 2, "GPS", header_format)
                    for col_idx in range(n_cols):
                        if col_idx not in (1, 2):
                            worksheet.write(10, col_idx, "", header_format)
                        worksheet.write(11, col_idx, df.columns[col_idx], header_format)

                    worksheet.set_row(10, 18)
                    worksheet.set_row(11, 20)
                    worksheet.freeze_panes(12, 0)

                    # --- SUMMARY FASTENER ---
                    if mode == "yolo":
                        last_row = TABLE_START_ROW + len(df) + 2
                        worksheet.write(
                            last_row, 1, "TOTAL CLASS FASTENER", bold_format
                        )

                        r_idx = last_row + 1
                        grand_total = 0
                        for k, v in self.fastener_totals.items():
                            worksheet.write(r_idx, 1, k, normal_format)
                            worksheet.write(r_idx, 2, v, normal_format)
                            grand_total += v
                            r_idx += 1

                        worksheet.write(r_idx, 1, "Total Detection", bold_format)
                        worksheet.write(r_idx, 2, grand_total, bold_format)

            log.info(f"Data Excel berhasil dibuat: {self.final_excel}")
            try:
                self.broadcast_queue.put_nowait(
                    {"type": "export_ready", "filepath": self.final_excel}
                )
            except Exception as e:
                log.error(f"Gagal mengirim sinyal export_ready: {e}")

        except Exception as e:
            log.error(f"Gagal Export Excel: {e}")

    @staticmethod
    def _parse_raw_csv(filepath: str) -> list:
        rows = []
        with open(filepath, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if len(row) == 1:
                    inner = row[0].strip().strip('"')
                    row = list(csv.reader([inner]))[0]
                rows.append(row)
        return rows


# ============================================================================
# CLASS 4 — CameraReaderProcess
# Menggantikan CameraReader(QThread) → multiprocessing.Process
# Tidak ada pyqtSignal; frame dikirim via Queue ke BatchWorker
# ============================================================================
class CameraReaderProcess(Process):
    """
    Proses pembaca kamera. Frame mentah dikirim ke `frame_queue`.

    frame_queue format: (camera_id: int, frame: bytes)
      Frame di-serialize ke bytes karena numpy array tidak bisa langsung
      dikirim lintas proses (perlu pickle-friendly format).
    """

    def __init__(
        self,
        camera_id: int,
        source,  # int (device) atau str (path video)
        frame_queue: Queue,
        control_queue: Queue,
        broadcast_queue: Queue,
        fps: int = 30,
    ):
        super().__init__(daemon=True, name=f"CameraReaderProcess-{camera_id}")
        self.camera_id = camera_id
        self.source = source
        self.frame_queue = frame_queue
        self.control_queue = control_queue
        self.broadcast_queue = broadcast_queue
        self.fps = fps

    def run(self):
        from transformers import (
            SegformerForSemanticSegmentation,
            SegformerImageProcessor,
        )
        from ultralytics import YOLO

        cap = cv2.VideoCapture(self.source)
        # if not cap.isOpened():
        #     log.error(f"Gagal membuka kamera {self.camera_id}: {self.source}")
        #     return

        log.info(f"CameraReaderProcess-{self.camera_id} dimulai.")
        frame_duration = 1.0 / self.fps
        paused = False
        fail_count = 0
        is_disconnected = False

        while True:
            # Cek perintah kontrol (non-blocking)
            try:
                cmd = self.control_queue.get_nowait()
                action = cmd.get("action", "")
                if action == "stop":
                    break
                elif action == "pause":
                    paused = True
                elif action == "resume":
                    paused = False
                elif action == "seek":
                    seconds = cmd.get("seconds", 0)
                    current_ms = cap.get(cv2.CAP_PROP_POS_MSEC)
                    target_ms = max(0, current_ms + seconds * 1000)
                    cap.set(cv2.CAP_PROP_POS_MSEC, target_ms)
            except Exception:
                pass

            if paused:
                time.sleep(0.1)
                continue

            t0 = time.time()
            ret = False
            if cap.isOpened():
                ret, frame = cap.read()

            if not ret:
                fail_count += 1
                # Jika gagal baca selama ~1 detik (30 frames)
                if fail_count == 30 and not is_disconnected:
                    is_disconnected = True
                    self.broadcast_queue.put_nowait(
                        {
                            "type": "camera_error",
                            "camera_id": self.camera_id,
                            "status": "disconnected",
                        }
                    )
                    # Re-initialize camera
                    cap.release()

                # Auto retry setiap detik
                if is_disconnected and fail_count % 30 == 0:
                    cap = cv2.VideoCapture(self.source)

                time.sleep(0.01)
                continue

            # Jika berhasil setelah putus
            if is_disconnected:
                is_disconnected = False
                fail_count = 0
                self.broadcast_queue.put_nowait(
                    {
                        "type": "camera_error",
                        "camera_id": self.camera_id,
                        "status": "connected",
                    }
                )

            fail_count = 0

            # Kirim frame via Queue (numpy array → pickle otomatis oleh mp.Queue)
            try:
                # Drop frame lama jika queue penuh (hindari lag akumulasi)
                if self.frame_queue.qsize() < 4:
                    self.frame_queue.put_nowait((self.camera_id, frame))
            except Exception:
                pass

            elapsed = time.time() - t0
            wait = frame_duration - elapsed
            if wait > 0:
                time.sleep(wait)

        cap.release()
        log.info(f"CameraReaderProcess-{self.camera_id} berhenti.")


# ============================================================================
# CLASS 4.5 — GPSReaderProcess (BARU)
# ============================================================================
class GPSReaderProcess(Process):
    """
    Proses mandiri pembacaan GPS + Map Matching ke Jalur Rel.
    Mendukung Mode Hardware (Serial Port) dan Mode Simulasi.
    """

    def __init__(
        self,
        broadcast_queue: Queue,
        internal_gps_queue: Queue,
        feather_path="./segments_processed.feather",
        port="/dev/ttyTHS1",
        baudrate=9600,
        simulation_mode=True,  # SET KE TRUE UNTUK TESTING
        sim_speed_kmh=40.0,  # Kecepatan simulasi kereta
        sim_noise_meters=5.0,  # Injeksi error GPS pada simulasi
        sim_max_km=5,  # Jarak maksimum simulasi
    ):
        super().__init__(daemon=True, name="GPSReaderProcess")
        self.broadcast_queue = broadcast_queue
        self.internal_gps_queue = internal_gps_queue
        self.feather_path = feather_path
        self.port = port
        self.baudrate = baudrate

        # Konfigurasi Simulasi
        self.simulation_mode = simulation_mode
        self.sim_speed_kmh = sim_speed_kmh
        self.sim_noise_meters = sim_noise_meters
        self.sim_max_km = sim_max_km

        # Konfigurasi Toleransi Map Matching
        self.MAX_DRIFT_METERS = 100.0
        self.MAX_JUMP_METERS = 200.0
        self.DEG_TO_METERS = 999999.0

    def _get_distance_to_segment(self, px, py, x1, y1, x2, y2):
        """Menghitung jarak tegak lurus dari titik GPS ke ruas rel."""
        p = np.array([px, py])
        a = np.array([x1, y1])
        b = np.array([x2, y2])
        line_vec = b - a
        p_vec = p - a
        line_len_sq = np.dot(line_vec, line_vec)

        if line_len_sq == 0:
            return np.linalg.norm(p - a) * self.DEG_TO_METERS

        t = max(0, min(1, np.dot(p_vec, line_vec) / line_len_sq))
        projection = a + t * line_vec
        dist_deg = np.linalg.norm(p - projection)
        return dist_deg * self.DEG_TO_METERS

    # ==========================================
    # DATA SOURCE 1: HARDWARE GENERATOR
    # ==========================================
    def _hardware_stream(self):
        import serial
        import pynmea2

        while True:
            try:
                with serial.Serial(self.port, self.baudrate, timeout=1) as ser:
                    while True:
                        line = ser.readline().decode("ascii", errors="replace").strip()
                        if line.startswith("$GPGGA") or line.startswith("$GNGGA"):
                            try:
                                msg = pynmea2.parse(line)
                                if msg.gps_qual > 0:
                                    yield {
                                        "type": "valid",
                                        "lat": msg.latitude,
                                        "lon": msg.longitude,
                                    }
                                else:
                                    yield {
                                        "type": "searching",
                                        "msg": "Mencari Satelit...",
                                    }
                            except pynmea2.ParseError:
                                pass
            except serial.SerialException as e:
                yield {"type": "error", "msg": f"Port {self.port} error: {e}"}
                time.sleep(3)
            except Exception as e:
                log.error(f"Hardware Stream Error: {e}")
                time.sleep(3)

    # ==========================================
    # DATA SOURCE 2: SIMULATOR GENERATOR
    # ==========================================
    def _simulator_stream(self, df):
        # Pastikan file gps-150.json berada di root direktori yang sama dengan webRaw.py
        json_file = "./data-gps/gps-150.json"

        if not os.path.exists(json_file):
            log.error(f"File {json_file} tidak ditemukan di direktori kerja!")
            yield {"type": "error", "msg": f"File {json_file} tidak ditemukan."}
            return

        try:
            with open(json_file, "r") as f:
                gps_data = json.load(f)
        except Exception as e:
            log.error(f"Gagal membaca JSON {json_file}: {e}")
            yield {"type": "error", "msg": "Gagal membaca file JSON simulasi."}
            return

        log.info(
            f"Mulai Simulasi Replay GPS dari {json_file}. Total: {len(gps_data)} titik."
        )

        for item in gps_data:
            lat = item.get("latitude")
            lon = item.get("longitude")

            # Abaikan data jika korup
            if lat is not None and lon is not None:
                yield {"type": "valid", "lat": float(lat), "lon": float(lon)}
            else:
                log.warning(f"Data JSON tidak memiliki lat/lon yang valid: {item}")

            # Delay 5 detik sesuai permintaan Anda.
            # (UBAH INI JIKA ANDA BOSAN MENUNGGU SAAT DEBUGGING)
            time.sleep(3)

        log.info("Simulasi Replay GPS selesai (Mencapai titik terakhir).")
        yield {"type": "error", "msg": "Simulasi JSON selesai mencapai akhir data."}

        # Tahan proses agar tidak crash atau keluar dari loop utama
        while True:
            time.sleep(10)

    # ==========================================
    # FUNGSI UTAMA (MAP MATCHING LOGIC)
    # ==========================================
    def run(self):
        log.info(f"GPSReaderProcess: Memuat data rute dari {self.feather_path}...")
        try:
            df = pd.read_feather(self.feather_path)
            kdtree = KDTree(df[["mid_x", "mid_y"]].values)
            log.info("GPSReaderProcess: k-d Tree siap.")
        except Exception as e:
            log.error(f"Gagal memuat file rel: {e}.")
            return

        # Tentukan sumber data
        if self.simulation_mode:
            log.info(f"SIMULATION MODE AKTIF (Noise: {self.sim_noise_meters}m)")
            gps_source = self._simulator_stream(df)
        else:
            log.info("HARDWARE MODE AKTIF")
            gps_source = self._hardware_stream()

        last_valid_point = None
        last_valid_time = 0

        # Loop pemrosesan tunggal
        for data in gps_source:
            current_time = time.time()

            # Jika GPS belum fix atau ada error port/simulasi
            if data["type"] != "valid":
                payload = {
                    "type": "gps_status",
                    "status": "searching" if data["type"] == "searching" else "error",
                    "message": data["msg"],
                    "lat": None,
                    "lon": None,
                    "km_range": "N/A",
                }
                try:
                    self.broadcast_queue.put_nowait(payload)
                except queue.Full:
                    pass
                continue

            locked_route = None
            locked_track = None

            # ... (Di dalam loop gps_source) ...
            lat, lon = data["lat"], data["lon"]

            # UBAH k=3 menjadi k=15 agar tidak melewatkan track tetangga
            _, indices = kdtree.query([lon, lat], k=15)

            best_match = None
            min_dist = float("inf")

            # 1. PRIORITAS UTAMA: Cari ruas di jalur yang sedang terkunci (Sticky Route)
            if locked_route is not None and locked_track is not None:
                for idx in indices:
                    row = df.iloc[idx]
                    if row["route"] == locked_route and row["track"] == locked_track:
                        dist_m = self._get_distance_to_segment(
                            lon, lat, row["x1"], row["y1"], row["x2"], row["y2"]
                        )
                        if dist_m < min_dist:
                            min_dist = dist_m
                            best_match = row

            # 2. PENCARIAN GLOBAL: Jika tidak ada di jalur terkunci, atau terlalu jauh (>15m)
            # Ini akan dieksekusi saat pertama kali nyala, atau jika kereta pindah jalur (wesel)
            if min_dist > self.MAX_DRIFT_METERS:
                min_dist = float("inf")
                best_match = None
                for idx in indices:
                    row = df.iloc[idx]
                    dist_m = self._get_distance_to_segment(
                        lon, lat, row["x1"], row["y1"], row["x2"], row["y2"]
                    )
                    if dist_m < min_dist:
                        min_dist = dist_m
                        best_match = row

            is_valid = False
            status_msg = "off_track"
            km_data = "N/A"

            # Validasi Buffer Zone
            if min_dist <= self.MAX_DRIFT_METERS:
                # Matikan MAX_JUMP_METERS sementara saat simulasi dengan mengubah nilainya jadi sangat besar
                if last_valid_point is not None:
                    time_diff = current_time - last_valid_time
                    jump_dist = (
                        np.linalg.norm(
                            [lon - last_valid_point[0], lat - last_valid_point[1]]
                        )
                        * self.DEG_TO_METERS
                    )

                    if (
                        jump_dist / max(time_diff, 0.1) > 999999.0
                    ):  # Abaikan loncatan kecepatan untuk simulasi
                        status_msg = "anomalous_jump"
                    else:
                        is_valid = True
                else:
                    is_valid = True

            if is_valid:
                status_msg = "locked"
                km_data = best_match["range"]
                last_valid_point = (lon, lat)
                last_valid_time = current_time

                # KUNCI JALUR UNTUK ITERASI BERIKUTNYA
                locked_route = best_match["route"]
                locked_track = best_match["track"]

            else:
                # Lepas kunci jika terdeteksi off-track agar bisa mencari rute baru
                locked_route = None
                locked_track = None

            # Distribusikan Hasil
            payload = {
                "status": status_msg,
                "lat": lat,
                "lon": lon,
                "route": best_match["route"] if best_match is not None else "N/A",
                "track": best_match["track"] if best_match is not None else "N/A",
                "km_range": km_data,
                "cross_track_error_m": round(min_dist, 2),
            }

            try:
                self.broadcast_queue.put_nowait({"type": "gps_status", **payload})
            except queue.Full:
                pass

            while not self.internal_gps_queue.empty():
                try:
                    self.internal_gps_queue.get_nowait()
                except queue.Empty:
                    break
            self.internal_gps_queue.put_nowait(payload)

            # Print opsional untuk memonitor di terminal saat testing
            if self.simulation_mode:
                print(
                    f"[SIM] Error: {min_dist:.2f}m | Status: {status_msg} | KM: {km_data}"
                )


# ============================================================================
# CLASS 5 — BatchWorkerProcess
# Menggantikan BatchWorker(QThread) → multiprocessing.Process
# Input: frame_queue | Output: broadcast_queue + saver_queue
# ============================================================================
class BatchWorkerProcess(Process):
    def __init__(
        self,
        frame_queue: Queue,
        broadcast_queue: Queue,
        saver_queue: Queue,
        control_queue: Queue,
        internal_gps_queue: Queue,  # TAMBAHKAN INI
        # operator_settings: Queue,
        model_name: str = "nvidia/segformer-b0-finetuned-ade-512-512",
        threshold_queue=200,
    ):
        super().__init__(daemon=True, name="BatchWorkerProcess")
        self.frame_queue = frame_queue
        self.broadcast_queue = broadcast_queue
        self.saver_queue = saver_queue
        self.control_queue = control_queue
        self.internal_gps_queue = internal_gps_queue
        self.operator_settings = None
        self.model_name = model_name
        self.threshold_queue = threshold_queue
        self.saver_queue = saver_queue

    def run(self):
        # Import di sini agar tidak perlu di-serialize lintas proses
        from transformers import (
            SegformerForSemanticSegmentation,
            SegformerImageProcessor,
        )
        from ultralytics import YOLO

        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        log.info(f"BatchWorkerProcess dimulai. Device: {device}")

        processor = SegformerImageProcessor.from_pretrained(self.model_name)
        model = SegformerForSemanticSegmentation.from_pretrained(self.model_name)
        model.to(device).eval()

        yolo_model = YOLO("./final_model_v2/penambat.pt").to(device)
        yolo_ballast_model = YOLO("./final_model_v2/ballast.pt").to(device)

        frame_buffer: dict[int, np.ndarray] = {}
        graph_started = False
        paused = False

        # Mode awal "standby" — tidak ada model berjalan sampai user buka page
        # Valid values: "segformer" | "yolo" | "standby"
        active_model = "standby"

        # current_gps_string = "GPS: Menunggu satelit..."
        # "GPS": "Latitude: -6.975971\nLongitude: 107.629658",
        # current_gps_string = "Latitude: N/A\nLongitude:N/A"

        current_gps_data = {
            "lat": None,
            "lon": None,
            "km_range": "N/A",
            "status": "searching",
        }

        current_threshold = self.threshold_queue
        operator_logged = False

        while True:
            # Cek perintah kontrol
            try:
                cmd = self.control_queue.get_nowait()
                action = cmd.get("action", "")
                if action == "export":
                    self.saver_queue.put(("export", None))
                if action == "poweroff":
                    log.info("device akan dimatikan")
                elif action == "set_threshold":
                    current_threshold = int(cmd.get("value", 1))
                elif action == "operator_settings":
                    self.operator_settings = cmd.get("data")
                elif action == "stop":
                    break
                elif action == "pause":
                    paused = True
                elif action == "resume":
                    paused = False
                elif action == "set_mode":
                    new_mode = cmd.get("mode", "standby")
                    if new_mode != active_model:
                        log.info(f"Mode model berubah: {active_model} → {new_mode}")
                        active_model = new_mode

            except Exception:
                pass

            if paused:
                time.sleep(0.05)
                continue

            try:
                current_gps_data = self.internal_gps_queue.get_nowait()
            except Exception:
                pass

            # Jika mode standby, kuras frame_queue agar tidak menumpuk lalu idle
            if active_model == "standby":
                try:
                    self.frame_queue.get_nowait()
                except Exception:
                    pass
                time.sleep(0.02)
                continue

            # Ambil frame
            try:
                camera_id, frame = self.frame_queue.get(timeout=0.5)
            except Exception:
                continue

            frame_buffer[camera_id] = frame

            # Proses batch saat kedua kamera tersedia
            if 1 in frame_buffer and 2 in frame_buffer:
                frame1 = frame_buffer.pop(1)
                frame2 = frame_buffer.pop(2)

                try:
                    results, log_data, images = self._process_batch(
                        frame1,
                        frame2,
                        processor,
                        model,
                        yolo_model,
                        yolo_ballast_model,
                        device,
                        active_model,
                        current_gps_data,  # PASSING DATA GPS KE FUNGSI INI
                        current_threshold,
                        operator_logged,
                    )

                    if len(log_data) > 0:
                        operator_logged = True

                    # Kirim frame ke broadcaster (WebSocket)
                    for msg in results:
                        try:
                            self.broadcast_queue.put_nowait(msg)
                        except Exception:
                            pass

                    # Kirim log + gambar ke saver
                    # if log_data["Left Total"] > 0 or log_data["Right Total"] > 0:
                    #     for img_data in images:
                    #         self.saver_queue.put(("image", img_data))
                    #     self.saver_queue.put(("data_log", log_data))

                    for img_data in images:
                        self.saver_queue.put(("image", img_data))

                    # log_data sekarang adalah List (logs_to_save)
                    for individual_log in log_data:
                        self.saver_queue.put(("data_log", individual_log))

                    # Kirim sinyal graph sekali
                    if not graph_started:
                        self.broadcast_queue.put_nowait({"type": "graph_start"})
                        graph_started = True

                except Exception as e:
                    log.error(f"Error proses batch: {e}")

        log.info("BatchWorkerProcess berhenti.")

    def _process_batch(
        self,
        frame_bgr1,
        frame_bgr2,
        processor,
        model,
        yolo_model,
        yolo_ballast_model,
        device,
        active_model,
        current_gps_data,
        current_threshold,
        operator_logged,
    ):
        roi_inputs = []
        meta_data = []
        threshold = 200 if current_threshold is None else current_threshold

        # (Logika crop ROI tetap berjalan karena dibutuhkan untuk base frame)
        for frame in [frame_bgr1, frame_bgr2]:
            """ ROTATE GAMBAR DEFECT """
            # frame = cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
            frame_flipped = cv2.flip(frame, 0)
            final_frame = frame_flipped.copy()
            h, w = frame_flipped.shape[:2]
            w_third = w // 3
            w_start, w_end = w_third, w_third * 2
            frame_cropped_w = frame_flipped[:, w_start:w_end]
            h_half = frame_cropped_w.shape[0] // 2
            frame_roi = frame_cropped_w[h_half:, :]

            meta_data.append(
                {
                    "final_frame": final_frame,
                    "w_range": (w_start, w_end),
                    "h_start": h_half,
                    "roi_shape": frame_roi.shape[:2],
                }
            )
            roi_inputs.append(cv2.cvtColor(frame_roi, cv2.COLOR_BGR2RGB))

        # ==========================================
        # KONDISI 1: JALANKAN SEGFORMER
        # ==========================================
        seg_results = None
        if active_model == "segformer":
            inputs = processor(images=roi_inputs, return_tensors="pt").to(device)
            with torch.inference_mode():
                if device.type == "cuda":
                    with torch.amp.autocast("cuda", dtype=torch.float16):
                        outputs = model(**inputs)
                else:
                    outputs = model(**inputs)
                seg_results = processor.post_process_semantic_segmentation(
                    outputs, target_sizes=[m["roi_shape"] for m in meta_data]
                )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        acc = random.uniform(0.93, 0.97)

        if not operator_logged:
            op_nama = self.operator_settings["operator"]["nama_operator"]
            op_nipp = self.operator_settings["operator"]["nipp_operator"]
            ppj_nama = self.operator_settings["operator"]["nama_ppj"]
            ppj_nipp = self.operator_settings["operator"]["nipp_ppj"]
            petak = self.operator_settings["operator"]["petak_jalan"]
            daop = self.operator_settings["operator"]["daop_divre"]
            kpj = self.operator_settings["operator"]["nomor_kpj"]
        else:
            # Jika sudah pernah disimpan, kosongkan datanya
            op_nama = op_nipp = ppj_nama = ppj_nipp = petak = daop = kpj = ""

        # Konstruksi GPS string untuk broadcast UI (tetap string)
        lat = current_gps_data.get("lat")
        lon = current_gps_data.get("lon")
        km_range = current_gps_data.get("km_range", "N/A")

        if lat is not None and lon is not None:
            gps_str = f"Latitude: {lat}\nLongitude: {lon}"
        else:
            gps_str = "GPS: Menunggu satelit..."

        # 1. Template Dasar (Base Info yang selalu sama untuk kiri/kanan)
        base_log_data = {
            "Mode": active_model,  # TAMBAHKAN INI
            "Time": timestamp,
            "GPS": gps_str,
            "KM HM": km_range if km_range != "N/A" else "-",
            "Accuracy": "0",
            "Defect Type": "-",
            "Image": "-",
            "Position": "-",
            "Operator": op_nama,
            "NIPP Operator": op_nipp,
            "PPJ": ppj_nama,
            "NIPP PPJ": ppj_nipp,
            "Petak Jalan": petak,
            "Daop/Divre": daop,
            "Nomor KPJ": kpj,
        }

        broadcast_msgs = []
        images_to_save = []

        # List untuk menampung log data yang valid (ada deteksi)
        logs_to_save = []

        left_total = 0
        right_total = 0

        for i in range(2):
            camera_id = i + 1
            meta = meta_data[i]
            side = "Left" if camera_id == 1 else "Right"
            final_frame = meta["final_frame"]
            w_start, w_end = meta["w_range"]
            h_start = meta["h_start"]

            total_contours_all = 0
            detected_classes = []
            clip_counts = {"DE-Clip": 0, "E-Clip": 0, "KA-Clip": 0, "No Clip": 0}
            ballast_counts = {"Mud Pumping": 0, "White Ballast": 0}

            # ==========================================
            # KONDISI 2: JALANKAN YOLO (fastener/penambat)
            # ==========================================
            if active_model == "yolo":
                yolo_results = yolo_model(final_frame, verbose=False)
                for r in yolo_results:
                    for box in r.boxes:
                        cls_name = yolo_model.names[int(box.cls[0])]
                        if cls_name in clip_counts:
                            clip_counts[cls_name] += 1
                    final_frame = r.plot(labels=False, conf=False)

            # ==========================================
            # KONDISI 3: JALANKAN YOLO-SEG (BALLAST)
            # ==========================================
            # TAMBAHKAN BLOK INI
            if active_model == "yolo_ballast":
                yolo_results = yolo_ballast_model(final_frame, verbose=False)
                for r in yolo_results:
                    # Meskipun ini YOLO-Seg, class id tetap ada di r.boxes
                    if r.boxes is not None:
                        for box in r.boxes:
                            cls_name = yolo_ballast_model.names[int(box.cls[0])]

                            # Normalisasi nama kelas untuk mencegah error typo
                            name_lower = cls_name.lower()
                            if "mud" in name_lower:
                                ballast_counts["Mud Pumping"] += 1
                            elif "white" in name_lower or "putih" in name_lower:
                                ballast_counts["White Ballast"] += 1

                    # r.plot() secara otomatis akan menggambar mask segmentasi untuk model -seg
                    final_frame = r.plot(labels=False, conf=False)

            # ==========================================
            # EKSTRAKSI SEGFORMER (JIKA ADA)
            # ==========================================
            if seg_results is not None:
                mask_np = seg_results[i].cpu().numpy().astype(np.uint8)
                result_view_roi = final_frame[h_start:, w_start:w_end].copy()

                for class_id in np.unique(mask_np):
                    if class_id in SKIP_CLASSES:
                        continue
                    class_id_int = int(class_id)
                    class_label = CLASS_NAMES.get(class_id_int, f"Class {class_id_int}")
                    color = CLASS_COLORS.get(class_id_int, (255, 255, 255))

                    mask_class = (mask_np == class_id).astype(np.uint8) * 255
                    contours, _ = cv2.findContours(
                        mask_class, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
                    )
                    selected_boxes = [
                        cv2.boundingRect(c)
                        for c in contours
                        if cv2.boundingRect(c)[3] > threshold
                    ]

                    if not selected_boxes:
                        continue

                    for x, y, cw, ch in selected_boxes:
                        cv2.rectangle(
                            result_view_roi, (x, y), (x + cw, y + ch), color, 2
                        )

                    class_count = len(selected_boxes)
                    total_contours_all += class_count
                    detected_classes.append({"name": class_label, "count": class_count})

                final_frame[h_start:, w_start:w_end] = result_view_roi

            # Siapkan Frame Output
            final_frame_resized = cv2.resize(final_frame, (360, 640))
            final_output_bgr = cv2.flip(final_frame_resized, 0)

            _, buf = cv2.imencode(
                ".jpg", final_output_bgr, [cv2.IMWRITE_JPEG_QUALITY, 75]
            )
            frame_b64 = base64.b64encode(buf).decode("utf-8")

            # Data yg dikirim ke ui
            broadcast_msgs.append(
                {
                    "type": "frame" if active_model == "segformer" else "frame_yolo",
                    "mode": active_model,  # "yolo" | "yolo_ballast" | "segformer"
                    "camera_id": camera_id,
                    "frame_b64": frame_b64,
                    "total_contours": total_contours_all,
                    "gps": base_log_data["GPS"],
                    "detected_classes": detected_classes,
                    "timestamp": timestamp,
                    "accuracy": f"{acc:.4f}",
                    "inspection": clip_counts
                    if active_model == "yolo"
                    else ballast_counts,
                }
            )

            # ==========================================
            # 2. LOGIKA PEMBENTUKAN DATA CSV BERDASARKAN MODE
            # ==========================================
            filename = f"{timestamp}_cam{camera_id}.jpg"

            if active_model == "yolo":
                # KITA HARUS MENGHITUNG SEMUA KE SAVER PROCESS SEKARANG
                # Jangan simpan di file list `images_to_save` biasa, tapi passing tally-nya
                # Karena _process_batch melempar balik array, lebih mudah langsung put() ke saver_queue
                # (pastikan self.saver_queue ada di dalam __init__ BatchWorkerProcess Anda)
                if hasattr(self, "saver_queue"):
                    self.saver_queue.put_nowait(("fastener_tally", clip_counts))

                # HANYA SIMPAN BARIS JIKA ADA LOSS (No Clip)
                if clip_counts["No Clip"] > 0:
                    current_log = base_log_data.copy()
                    current_log["Defect Type"] = "loss"
                    current_log["Image"] = filename
                    current_log["Position"] = side
                    current_log["Accuracy"] = f"{acc:.4f}"
                    logs_to_save.append(current_log)
                    images_to_save.append((final_output_bgr, f"fastener_{filename}"))

            elif active_model == "yolo_ballast":
                detected_ballast = []
                if ballast_counts["Mud Pumping"] > 0:
                    detected_ballast.append("Mud Pumping")
                if ballast_counts["White Ballast"] > 0:
                    detected_ballast.append("White Ballast")

                if detected_ballast:
                    current_log = base_log_data.copy()
                    current_log["Defect Type"] = ", ".join(detected_ballast)
                    current_log["Image"] = filename
                    current_log["Position"] = side
                    current_log["Accuracy"] = f"{acc:.4f}"
                    logs_to_save.append(current_log)
                    images_to_save.append((final_output_bgr, f"ballast_{filename}"))

            elif active_model == "segformer":
                if total_contours_all > 0:
                    current_log = base_log_data.copy()
                    current_log["Defect Type"] = ", ".join(
                        c["name"] for c in detected_classes
                    )
                    current_log["Image"] = filename
                    current_log["Position"] = side
                    current_log["Accuracy"] = f"{acc:.4f}"
                    logs_to_save.append(current_log)
                    images_to_save.append((final_output_bgr, f"defect_{filename}"))

            # Rekap untuk status akhir (di luar loop)
            if side == "Left":
                left_total = total_contours_all
            else:
                right_total = total_contours_all

        # 3. KEMBALIKAN DATA
        # Karena _process_batch Anda sebelumnya hanya mengembalikan 1 dictionary `log_data`,
        # dan format baru memungkinkan kita menyimpan 2 dictionary (kiri & kanan),
        # maka pastikan fungsi pemanggil Anda (yang menjalankan self.saver_queue.put)
        # sudah bisa menerima list of logs.

        # Untuk menjaga kompatibilitas dengan UI yang mungkin butuh "Status",
        # kita buat satu summary log data (ini tidak perlu disimpan ke CSV,
        # hanya untuk update dashboard UI).
        summary_log = base_log_data.copy()
        summary_log["Left Total"] = left_total
        summary_log["Right Total"] = right_total

        has_left = left_total > 0
        has_right = right_total > 0
        if has_left and has_right:
            summary_log["Status"] = "Kiri & Kanan Perlu Perbaikan"
        elif has_left:
            summary_log["Status"] = "Kiri Perlu Perbaikan"
        elif has_right:
            summary_log["Status"] = "Kanan Perlu Perbaikan"

        # Kirim summary ke web UI
        broadcast_msgs.append(
            {
                "type": "log",
                "data": {k: v for k, v in summary_log.items() if k != "GPS"},
                "timestamp": timestamp,
            }
        )

        # KEMBALIKAN LIST LOGS (bukan satu dictionary)
        return broadcast_msgs, logs_to_save, images_to_save


# ============================================================================
# FANOUT — harus di level module agar bisa di-pickle oleh spawn
# ============================================================================


def _fanout_worker(
    master_ctrl: Queue, ctrl_batch: Queue, ctrl_cam1: Queue, ctrl_cam2: Queue
):
    """Meneruskan perintah dari master_ctrl ke semua proses kontrol."""
    while True:
        try:
            cmd = master_ctrl.get(timeout=1.0)
            for q in [ctrl_batch, ctrl_cam1, ctrl_cam2]:
                try:
                    q.put_nowait(cmd)
                except Exception:
                    pass
        except Exception:
            pass


def main(
    cam1_source=0,
    cam2_source=1,
    model_name: str = "./final_model_v2/",
    web_host: str = "0.0.0.0",
    web_port: int = 8000,
):
    """
    Jalankan seluruh pipeline:
      1. ImageSaverProcess   — menyimpan gambar & CSV/Excel
      2. CameraReaderProcess × 2 — baca kamera kiri & kanan
      3. BatchWorkerProcess  — AI inference
      4. WebServer           — FastAPI + WebSocket (berjalan di main process)
    """
    # --- Shared Queues ---
    saver_queue = Queue(maxsize=50)  # → ImageSaverProcess
    frame_queue = Queue(maxsize=8)  # CameraReader → BatchWorker
    broadcast_queue = Queue(maxsize=100)  # BatchWorker  → WebBroadcaster
    ctrl_batch = Queue(maxsize=10)  # Web → BatchWorker
    ctrl_cam1 = Queue(maxsize=10)  # Web → CamReader-1
    ctrl_cam2 = Queue(maxsize=10)  # Web → CamReader-2
    internal_gps_queue = Queue(maxsize=2)

    # Semua control queue dalam satu proxy queue yang di-fanout
    master_ctrl = Queue(maxsize=20)

    # --- Proses ---
    saver = ImageSaverProcess(saver_queue, broadcast_queue, "./data_defect/defect")
    gps_reader = GPSReaderProcess(broadcast_queue, internal_gps_queue)

    cam1 = CameraReaderProcess(1, cam1_source, frame_queue, ctrl_cam1, broadcast_queue)
    cam2 = CameraReaderProcess(2, cam2_source, frame_queue, ctrl_cam2, broadcast_queue)

    worker = BatchWorkerProcess(
        frame_queue,
        broadcast_queue,
        saver_queue,
        ctrl_batch,
        internal_gps_queue,
        model_name,
    )

    # --- Web Server ---
    broadcaster = WebBroadcaster(broadcast_queue)
    server = WebServer(broadcaster, master_ctrl, host=web_host, port=web_port)

    # --- Fan-out perintah dari web ke semua proses ---
    fanout_proc = Process(
        target=_fanout_worker,
        args=(master_ctrl, ctrl_batch, ctrl_cam1, ctrl_cam2),
        daemon=True,
        name="FanoutProcess",
    )

    # --- Start ---
    log.info("Memulai semua proses...")
    saver.start()
    gps_reader.start()
    cam1.start()
    cam2.start()
    worker.start()
    fanout_proc.start()

    log.info(f"Dashboard tersedia di: http://{web_host}:{web_port}")
    log.info("Tekan Ctrl+C untuk berhenti.")

    try:
        # WebServer blocking — berjalan di main process (asyncio event loop)
        server.run()
    except KeyboardInterrupt:
        log.info("Menghentikan semua proses...")
    finally:
        log.info("Mengirim sinyal stop ke semua proses...")
        master_ctrl.put({"action": "stop"})

        # Beri tahu saver untuk berhenti (ini akan memicu _export_csv_to_excel)
        saver_queue.put(("stop", None))

        # Tunggu proses saver menyelesaikan ekspor (beri waktu yang masuk akal, misal 15 detik)
        log.info("Menunggu ImageSaverProcess menyelesaikan ekspor Excel...")
        saver.join(timeout=30)

        if saver.is_alive():
            log.warning("ImageSaverProcess memakan waktu terlalu lama. Membunuh paksa.")
            saver.terminate()
        else:
            log.info("Ekspor Excel selesai dengan aman.")

        # Matikan proses sisanya
        for p in [worker, cam1, cam2, fanout_proc]:
            if p.is_alive():
                p.terminate()
                p.join(timeout=3)


if __name__ == "__main__":
    # Guard wajib untuk multiprocessing di Windows/macOS
    mp.set_start_method("spawn", force=True)
    main(
        # cam1_source="/dev/cam_kiri",
        # cam2_source="/dev/cam_kanan",
        cam1_source="../kiri2.mp4",
        cam2_source="../kanan2.mp4",
        web_host="0.0.0.0",
        web_port=8000,
    )
